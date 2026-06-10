from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle


class ReportGenerator:
    def __init__(self, output_dir: str):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def export_excel(self, filename: str, sections: dict[str, list[dict]]) -> str:
        path = self.output_dir / filename
        wb = Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)

        header_fill = PatternFill("solid", fgColor="0F172A")
        header_font = Font(color="FFFFFF", bold=True)

        for sheet_name, rows in sections.items():
            ws = wb.create_sheet(title=sheet_name[:31])
            if not rows:
                ws.append(["Sin datos"])
                continue
            headers = list(rows[0].keys())
            ws.append(headers)
            for row in rows:
                ws.append([row.get(h) for h in headers])
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

        wb.save(path)
        return str(path)

    def export_pdf(self, filename: str, title: str, sections: dict[str, list[dict]]) -> str:
        path = self.output_dir / filename
        doc = SimpleDocTemplate(str(path), pagesize=landscape(A4), title=title)
        styles = getSampleStyleSheet()
        story = [Paragraph(title, styles["Title"]), Spacer(1, 12)]

        for section_name, rows in sections.items():
            story.append(Paragraph(section_name, styles["Heading2"]))
            story.append(Spacer(1, 6))
            if not rows:
                story.append(Paragraph("Sin datos disponibles.", styles["BodyText"]))
                story.append(Spacer(1, 10))
                continue
            headers = list(rows[0].keys())
            table_data = [headers]
            for row in rows:
                table_data.append([str(row.get(h, "")) for h in headers])
            table = Table(table_data, repeatRows=1)
            table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F172A")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]))
            story.append(table)
            story.append(Spacer(1, 14))

        doc.build(story)
        return str(path)

    @staticmethod
    def timestamped_name(prefix: str, extension: str) -> str:
        return f"{prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.{extension.lstrip('.')}"

